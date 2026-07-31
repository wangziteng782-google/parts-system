#!/usr/bin/env python
"""把 excel/new_parts.xlsx 的“补录”工作表导入 parts 主表。

默认只预览，不修改数据库、不上传图片：
    python jobs/import_new_parts.py

测试导入一条：
    python jobs/import_new_parts.py --execute --limit 1

后续全量导入：
    python jobs/import_new_parts.py --execute

脚本会在 Excel 同目录保存 new_parts_import_state.json。已经成功导入的行会自动跳过，
因此测试导入的一条不会在后续全量运行时重复写入。
"""

from __future__ import annotations

import argparse
import hashlib
import json
import mimetypes
import os
import re
import sys
import uuid
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any
from urllib.parse import quote

import pymysql
from openpyxl import load_workbook
from qiniu import Auth, put_data


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL_PATH = PROJECT_DIR / "excel" / "new_parts.xlsx"
DEFAULT_STATE_PATH = PROJECT_DIR / "excel" / "new_parts_import_state.json"
DEFAULT_ERROR_PATH = PROJECT_DIR / "excel" / "new_parts_import_errors.jsonl"
QINIU_CONFIG_PATH = PROJECT_DIR / "qiniu_config.local.json"
DEFAULT_SHEET_NAME = "补录"
MAX_IMAGE_SIZE = 10 * 1024 * 1024

# Excel 中重复出现“备注、填报人、更新时间”，因此这里按确定的列号映射。
# 列号为 Excel 的 1-based 列号；SKU 列明确忽略，product_type 在本批 Excel 中不存在。
CELL_COLUMN_MAP = {
    "sku_code": None,
    "product_name": 2,
    "product_brand": 3,
    "model": 4,
    "supplier": 5,
    "warranty": 6,
    "applicable_elevator_brand": 7,
    "nature": 8,
    "substitute_model": 9,
    "precautions": 10,
    "category": 11,
    "technical_params": 12,
    "purchase_cost": 13,
    "purchase_special_invoice": 14,
    "purchase_general_invoice": 15,
    "purchase_shipping": 16,
    "retail_price": 17,
    "retail_ladder_price": 18,
    "retail_tax": 19,
    "retail_shipping": 20,
    "remark": 21,
    "daily_cutoff_time": 22,
    "quote_validity": 23,
    "shipping_origin": 24,
    "shipping_time": 25,
    "remark_2": 26,
    "updater": 27,
    "filler": 39,
    "update_time": 40,
    "filler_2": 41,
    "update_time_2": 42,
    "filler_ip": 43,
    "product_type": None,
}

IMAGE_FIELDS = [
    "key_part_images",
    "actual_photos",
    "product_image_3",
    "product_image_4",
    "product_image_5",
    "product_image_6",
    "product_image_7",
    "product_image_8",
    "product_image_9",
    "product_image_10",
    "product_detail_images",
]

# 关键部位图片、关键部位图片2/3/4 全部合并到 key_part_images。
IMAGE_COLUMN_MAP = {
    28: "key_part_images",
    29: "actual_photos",
    30: "product_image_3",
    31: "product_image_4",
    32: "product_image_5",
    33: "product_image_6",
    34: "product_image_7",
    35: "product_image_8",
    36: "product_image_9",
    37: "product_image_10",
    38: "product_detail_images",
    44: "key_part_images",
    45: "key_part_images",
    46: "key_part_images",
}

EXPECTED_HEADERS = {
    2: "产品名称",
    3: "产品品牌",
    4: "型号",
    27: "更新人",
    28: "关键部位图片",
    29: "实物图照片",
    38: "商品详情图片",
    43: "填报ip",
    44: "关键部位图片2",
    45: "关键部位图片3",
    46: "关键部位图片4",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="将 new_parts.xlsx 导入 parts 主表")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_PATH, help="Excel 文件路径")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="工作表名称")
    parser.add_argument("--execute", action="store_true", help="实际上传图片并写入数据库；不加则只预览")
    parser.add_argument("--limit", type=int, default=None, help="本次最多处理多少条未导入数据")
    parser.add_argument("--start-row", type=int, default=None, help="指定数据开始行；默认自动识别表头")
    parser.add_argument("--state-file", type=Path, default=DEFAULT_STATE_PATH, help="断点及去重状态文件")
    parser.add_argument("--error-file", type=Path, default=DEFAULT_ERROR_PATH, help="错误记录文件")
    parser.add_argument("--stop-on-error", action="store_true", help="遇到一条错误后立即停止")
    return parser.parse_args()


def load_local_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as file:
        value = json.load(file)
    return value if isinstance(value, dict) else {}


def normalize_domain(value: str) -> str:
    domain = value.strip()
    markdown_match = re.fullmatch(r"\[[^\]]+\]\((https?://[^)]+)\)", domain)
    if markdown_match:
        domain = markdown_match.group(1)
    if domain and not domain.startswith(("http://", "https://")):
        domain = "https://" + domain
    return domain.rstrip("/")


def load_qiniu_config() -> dict[str, str]:
    local = load_local_json(QINIU_CONFIG_PATH)
    config = {
        "access_key": os.getenv("QINIU_ACCESS_KEY", "").strip() or str(local.get("access_key", "")).strip(),
        "secret_key": os.getenv("QINIU_SECRET_KEY", "").strip() or str(local.get("secret_key", "")).strip(),
        "bucket": os.getenv("QINIU_BUCKET", "").strip() or str(local.get("bucket", "")).strip(),
        "domain": normalize_domain(
            os.getenv("QINIU_DOMAIN", "").strip() or str(local.get("domain", "")).strip()
        ),
    }
    missing = [key for key, value in config.items() if not value]
    if missing:
        raise RuntimeError("七牛云配置不完整：" + "、".join(missing))
    return config


def load_db_config() -> dict[str, Any]:
    return {
        "host": os.getenv("PARTS_DB_HOST", "localhost"),
        "port": int(os.getenv("PARTS_DB_PORT", "3306")),
        "user": os.getenv("PARTS_DB_USER", "root"),
        "password": os.getenv("PARTS_DB_PASSWORD", "1234"),
        "database": os.getenv("PARTS_DB_NAME", "parts_database"),
        "charset": "utf8mb4",
        "cursorclass": pymysql.cursors.DictCursor,
        "autocommit": False,
    }


def source_fingerprint(path: Path, sheet_name: str) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        while chunk := file.read(4 * 1024 * 1024):
            digest.update(chunk)
    digest.update(sheet_name.encode("utf-8"))
    return digest.hexdigest()


def empty_state(fingerprint: str, excel_path: Path, sheet_name: str) -> dict[str, Any]:
    return {
        "version": 1,
        "source_fingerprint": fingerprint,
        "excel_path": str(excel_path.resolve()),
        "sheet": sheet_name,
        "rows": {},
        "uploaded_images": {},
    }


def load_state(path: Path, fingerprint: str, excel_path: Path, sheet_name: str) -> dict[str, Any]:
    if not path.exists():
        return empty_state(fingerprint, excel_path, sheet_name)
    state = load_local_json(path)
    state.setdefault("rows", {})
    state.setdefault("uploaded_images", {})
    return state


def save_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8") as file:
        json.dump(state, file, ensure_ascii=False, indent=2)
    os.replace(temp_path, path)


def append_error(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        file.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")


def acquire_lock(state_path: Path) -> tuple[int, Path]:
    lock_path = state_path.with_suffix(state_path.suffix + ".lock")
    try:
        descriptor = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as error:
        raise RuntimeError(f"检测到另一个导入任务正在运行，或上次异常退出：{lock_path}") from error
    os.write(descriptor, str(os.getpid()).encode("ascii"))
    return descriptor, lock_path


def release_lock(descriptor: int, lock_path: Path) -> None:
    os.close(descriptor)
    lock_path.unlink(missing_ok=True)


def find_header_row(worksheet) -> int:
    for row_number in range(1, min(worksheet.max_row, 30) + 1):
        if (
            clean_text(worksheet.cell(row_number, 2).value) == "产品名称"
            and clean_text(worksheet.cell(row_number, 4).value) == "型号"
            and clean_text(worksheet.cell(row_number, 28).value) == "关键部位图片"
        ):
            return row_number
    raise RuntimeError("未找到预期表头行（产品名称、型号、关键部位图片）")


def validate_headers(worksheet, header_row: int) -> None:
    errors = []
    for column, expected in EXPECTED_HEADERS.items():
        actual = clean_text(worksheet.cell(header_row, column).value)
        if actual.lower() != expected.lower():
            errors.append(f"第{column}列应为“{expected}”，实际为“{actual or '空'}”")
    if errors:
        raise RuntimeError("Excel 表头与脚本映射不一致：\n" + "\n".join(errors))


def reconcile_changed_source(
    state: dict[str, Any],
    worksheet,
    fingerprint: str,
    excel_path: Path,
    sheet_name: str,
    state_path: Path,
) -> dict[str, Any]:
    """Excel 变化后按产品名称+型号迁移已导入行，避免因排序变化重复导入。"""
    if state.get("source_fingerprint") == fingerprint:
        return state

    imported_rows = state.get("rows", {})
    if not imported_rows:
        state.update(
            {
                "source_fingerprint": fingerprint,
                "excel_path": str(excel_path.resolve()),
                "sheet": sheet_name,
            }
        )
        save_state(state_path, state)
        return state

    identity_rows: dict[tuple[str, str], list[int]] = defaultdict(list)
    for row_number in range(1, worksheet.max_row + 1):
        identity = (
            clean_text(worksheet.cell(row_number, 2).value),
            clean_text(worksheet.cell(row_number, 4).value),
        )
        if identity[0]:
            identity_rows[identity].append(row_number)

    migrated_rows: dict[str, Any] = {}
    migrations = []
    problems = []
    for old_row_text, record in imported_rows.items():
        old_row = int(old_row_text)
        identity = (
            clean_text(record.get("product_name")),
            clean_text(record.get("model")),
        )
        current_identity = (
            clean_text(worksheet.cell(old_row, 2).value),
            clean_text(worksheet.cell(old_row, 4).value),
        )
        if current_identity == identity:
            new_row = old_row
        else:
            matches = identity_rows.get(identity, [])
            if len(matches) == 1:
                new_row = matches[0]
            elif not matches:
                problems.append(
                    f"原Excel行{old_row}：{identity[0]} / {identity[1] or '无型号'}，"
                    "在新文件中找不到"
                )
                continue
            else:
                problems.append(
                    f"原Excel行{old_row}：{identity[0]} / {identity[1] or '无型号'}，"
                    f"在新文件中出现多次（行号：{matches}）"
                )
                continue

        new_key = str(new_row)
        if new_key in migrated_rows:
            problems.append(f"多条历史状态同时匹配到新Excel第{new_row}行")
            continue
        migrated_rows[new_key] = record
        if new_row != old_row:
            migrations.append((old_row, new_row, identity[0], identity[1]))

    if problems:
        raise RuntimeError(
            "Excel 文件内容或排序已变化，自动迁移导入状态时存在歧义：\n"
            + "\n".join(problems)
            + "\n请不要删除状态文件，先检查上述产品后再处理。"
        )

    state["rows"] = migrated_rows
    state["source_fingerprint"] = fingerprint
    state["excel_path"] = str(excel_path.resolve())
    state["sheet"] = sheet_name
    state["reconciled_at"] = datetime.now().isoformat(timespec="seconds")
    save_state(state_path, state)

    print("检测到 Excel 内容或排序变化，已安全核对历史导入记录：")
    if migrations:
        for old_row, new_row, product_name, model in migrations:
            print(
                f"  已导入产品 {product_name} / {model or '无型号'}："
                f"Excel行 {old_row} -> {new_row}"
            )
    else:
        print("  已导入产品仍在原行号，仅更新文件指纹。")
    return state


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def to_db_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def to_db_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    text = clean_text(value)
    if text:
        for pattern in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d",
        ):
            try:
                return datetime.strptime(text, pattern)
            except ValueError:
                continue
    # 新导入数据没有第二更新时间时，记录实际导入时间。
    return datetime.now()


def build_image_index(worksheet) -> dict[int, dict[str, list[Any]]]:
    result: dict[int, dict[str, list[Any]]] = defaultdict(lambda: defaultdict(list))
    for image in worksheet._images:
        anchor = getattr(image.anchor, "_from", None)
        if anchor is None:
            continue
        row_number = anchor.row + 1
        column_number = anchor.col + 1
        field = IMAGE_COLUMN_MAP.get(column_number)
        if field:
            result[row_number][field].append(image)
    return result


def image_bytes_and_type(image) -> tuple[bytes, str, str]:
    content = image._data()
    if not content:
        raise RuntimeError("Excel 嵌入图片内容为空")
    if len(content) > MAX_IMAGE_SIZE:
        raise RuntimeError(f"Excel 嵌入图片超过 10MB：{len(content) / 1024 / 1024:.2f}MB")

    image_format = (getattr(image, "format", None) or "").lower()
    if image_format == "jpg":
        image_format = "jpeg"
    extension_map = {
        "jpeg": ".jpg",
        "png": ".png",
        "gif": ".gif",
        "webp": ".webp",
        "bmp": ".bmp",
    }
    extension = extension_map.get(image_format)
    if not extension:
        guessed = mimetypes.guess_extension(getattr(image, "mime_type", "") or "")
        extension = guessed or ".jpg"
    mime_type = mimetypes.types_map.get(extension, "image/jpeg")
    return content, extension, mime_type


def upload_image(
    image,
    qiniu_config: dict[str, str],
    auth: Auth,
    state: dict[str, Any],
    state_path: Path,
) -> str:
    content, extension, mime_type = image_bytes_and_type(image)
    image_hash = hashlib.sha256(content).hexdigest()
    cached = state["uploaded_images"].get(image_hash)
    if cached and cached.get("url"):
        return cached["url"]

    key = f"{datetime.now():%Y%m%d}/{uuid.uuid4().hex}{extension}"
    token = auth.upload_token(qiniu_config["bucket"], key, 3600)
    result, info = put_data(token, key, content, mime_type=mime_type, check_crc=True)
    if not result or result.get("key") != key:
        qiniu_error = (
            getattr(info, "text_body", None)
            or getattr(info, "error", None)
            or "未知错误"
        )
        raise RuntimeError(f"七牛云上传失败：{qiniu_error}")

    url = f"{qiniu_config['domain']}/{quote(key, safe='/')}"
    state["uploaded_images"][image_hash] = {
        "key": key,
        "url": url,
        "uploaded_at": datetime.now().isoformat(timespec="seconds"),
    }
    # 图片上传成功后立即记录，避免数据库失败后重复上传同一张图。
    save_state(state_path, state)
    return url


def row_has_data(worksheet, row_number: int, image_index: dict[int, Any]) -> bool:
    text_columns = [column for column in CELL_COLUMN_MAP.values() if column]
    return any(worksheet.cell(row_number, column).value not in (None, "") for column in text_columns) or bool(
        image_index.get(row_number)
    )


def build_row_values(
    worksheet,
    row_number: int,
    row_images: dict[str, list[Any]],
    qiniu_config: dict[str, str] | None,
    auth: Auth | None,
    state: dict[str, Any],
    state_path: Path,
    execute: bool,
) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for field, column in CELL_COLUMN_MAP.items():
        if field == "update_time_2":
            cell_value = worksheet.cell(row_number, column).value if column else None
            values[field] = to_db_datetime(cell_value)
        elif column is None:
            values[field] = None
        else:
            values[field] = to_db_text(worksheet.cell(row_number, column).value)

    for image_field in IMAGE_FIELDS:
        images = row_images.get(image_field, [])
        if execute:
            urls = [
                upload_image(image, qiniu_config, auth, state, state_path)
                for image in images
            ]
            values[image_field] = json.dumps(urls, ensure_ascii=False) if urls else None
        else:
            values[image_field] = f"[{len(images)}张嵌入图片]" if images else None
    return values


def insert_part(connection, values: dict[str, Any]) -> int:
    fields = list(values)
    field_sql = ", ".join(f"`{field}`" for field in fields)
    placeholders = ", ".join(["%s"] * len(fields))
    sql = f"INSERT INTO parts ({field_sql}) VALUES ({placeholders})"
    with connection.cursor() as cursor:
        cursor.execute(sql, [values[field] for field in fields])
        return cursor.lastrowid


def print_preview(row_number: int, values: dict[str, Any]) -> None:
    preview = {
        "excel_row": row_number,
        "product_name": values.get("product_name"),
        "model": values.get("model"),
        "product_brand": values.get("product_brand"),
        "supplier": values.get("supplier"),
        "sku_code": values.get("sku_code"),
        "product_type": values.get("product_type"),
        "images": {field: values.get(field) for field in IMAGE_FIELDS if values.get(field)},
    }
    print(json.dumps(preview, ensure_ascii=False, default=str))


def main() -> int:
    args = parse_args()
    excel_path = args.excel.resolve()
    state_path = args.state_file.resolve()
    error_path = args.error_file.resolve()
    if not excel_path.exists():
        raise RuntimeError(f"Excel 文件不存在：{excel_path}")
    if args.limit is not None and args.limit <= 0:
        raise RuntimeError("--limit 必须大于 0")

    fingerprint = source_fingerprint(excel_path, args.sheet)
    state = load_state(state_path, fingerprint, excel_path, args.sheet)
    workbook = load_workbook(excel_path, data_only=True, read_only=False)
    if args.sheet not in workbook.sheetnames:
        workbook.close()
        raise RuntimeError(f"工作表不存在：{args.sheet}")
    worksheet = workbook[args.sheet]
    header_row = find_header_row(worksheet)
    validate_headers(worksheet, header_row)
    state = reconcile_changed_source(
        state,
        worksheet,
        fingerprint,
        excel_path,
        args.sheet,
        state_path,
    )
    data_start_row = args.start_row or (header_row + 1)
    image_index = build_image_index(worksheet)

    imported_rows = state["rows"]
    candidates = [
        row_number
        for row_number in range(data_start_row, worksheet.max_row + 1)
        if row_has_data(worksheet, row_number, image_index)
        and str(row_number) not in imported_rows
    ]
    if args.limit is not None:
        candidates = candidates[: args.limit]

    print(
        f"文件={excel_path.name} | 工作表={args.sheet} | 表头行={header_row} | "
        f"数据候选={len(candidates)} | 已导入跳过={len(imported_rows)} | "
        f"模式={'执行' if args.execute else '预览'}"
    )
    if not candidates:
        print("没有需要导入的数据。")
        workbook.close()
        return 0

    if not args.execute:
        for row_number in candidates[: min(len(candidates), 10)]:
            values = build_row_values(
                worksheet,
                row_number,
                image_index.get(row_number, {}),
                None,
                None,
                state,
                state_path,
                execute=False,
            )
            print_preview(row_number, values)
        if len(candidates) > 10:
            print(f"预览仅显示前10条，尚有 {len(candidates) - 10} 条。")
        print("预览完成；如确认无误，请加 --execute 执行。")
        workbook.close()
        return 0

    qiniu_config = load_qiniu_config()
    auth = Auth(qiniu_config["access_key"], qiniu_config["secret_key"])
    connection = pymysql.connect(**load_db_config())
    lock_descriptor, lock_path = acquire_lock(state_path)
    success_count = 0
    error_count = 0
    try:
        for position, row_number in enumerate(candidates, start=1):
            try:
                values = build_row_values(
                    worksheet,
                    row_number,
                    image_index.get(row_number, {}),
                    qiniu_config,
                    auth,
                    state,
                    state_path,
                    execute=True,
                )
                if not values.get("product_name"):
                    raise RuntimeError("产品名称为空，已拒绝写入")
                part_id = insert_part(connection, values)
                connection.commit()
                state["rows"][str(row_number)] = {
                    "part_id": part_id,
                    "product_name": values.get("product_name"),
                    "model": values.get("model"),
                    "imported_at": datetime.now().isoformat(timespec="seconds"),
                }
                save_state(state_path, state)
                success_count += 1
                print(
                    f"[{position}/{len(candidates)}] 成功 | Excel行={row_number} | "
                    f"parts.id={part_id} | {values.get('product_name')} | {values.get('model') or '无型号'}"
                )
            except Exception as error:
                connection.rollback()
                error_count += 1
                payload = {
                    "time": datetime.now().isoformat(timespec="seconds"),
                    "excel_row": row_number,
                    "product_name": to_db_text(worksheet.cell(row_number, 2).value),
                    "model": to_db_text(worksheet.cell(row_number, 4).value),
                    "error": str(error),
                }
                append_error(error_path, payload)
                print(
                    f"[{position}/{len(candidates)}] 失败 | Excel行={row_number} | {error}",
                    file=sys.stderr,
                )
                if args.stop_on_error:
                    break
    finally:
        release_lock(lock_descriptor, lock_path)
        connection.close()
        workbook.close()

    print(
        f"导入结束 | 成功={success_count} | 失败={error_count} | "
        f"状态文件={state_path}"
    )
    return 1 if error_count else 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        print("用户中断导入。", file=sys.stderr)
        raise SystemExit(130)
    except Exception as error:
        print(f"导入任务启动失败：{error}", file=sys.stderr)
        raise SystemExit(1)
